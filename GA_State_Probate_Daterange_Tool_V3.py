import re
import time
from urllib.parse import urljoin
from datetime import datetime, timedelta

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import Workbook


# ---------------- Selenium helpers ---------------- #

def safe_get_text_by_id(driver, element_id):
    """Return text of element by ID or empty string."""
    try:
        el = driver.find_element(By.ID, element_id)
        return el.text.strip()
    except Exception:
        return ""


def get_paging_info(driver):
    """
    Robust pager detection:
    - If Telerik 'rgInfoPart' exists => parse "X items in Y pages"
    - Else fallback to 1 page
    """
    total_records = None
    total_pages = 1

    info_elems = driver.find_elements(By.CSS_SELECTOR, "div.rgWrap.rgInfoPart")
    if info_elems:
        info_text = info_elems[0].text.strip()
        m = re.search(r"(\d+)\s+items\s+in\s+(\d+)\s+pages", info_text)
        if m:
            total_records = int(m.group(1))
            total_pages = int(m.group(2))

    # Fallback: infer pages from numeric pager links (if present)
    if total_pages == 1:
        num_links = driver.find_elements(By.CSS_SELECTOR, "div.rgWrap.rgNumPart a")
        nums = []
        for a in num_links:
            t = a.text.strip()
            if t.isdigit():
                nums.append(int(t))
        if nums:
            total_pages = max(nums)

    # Fallback: record count from current page rows (only if total_records unknown)
    if total_records is None:
        rows = driver.find_elements(
            By.CSS_SELECTOR,
            "#ctl00_cpMain_rgEstates_ctl00 tbody tr[id^='ctl00_cpMain_rgEstates_ctl00__']"
        )
        total_records = len(rows)

    return total_records, total_pages

# ---------------- NEW LOGIC ---------------- #

PRIORITY_PARTY = ["PETITIONER", "EXECUTOR", "ADMINISTRATOR"]

PRIORITY_FILINGS = [
    "Petition To Probate Will In Solemn Form",
    "Petition to Probate Will in Common Form",
    "Petition For Letters Of Administration",
    "Petition For Temporary Letters Of Administration",
    "Petition For Order Declaring No Administration Necessary"
]


def filter_and_sort_parties(types, names, addresses):
    filtered = []

    for t, n, a in zip(types, names, addresses):
        upper = t.upper()
        for p in PRIORITY_PARTY:
            if p in upper:
                filtered.append((p, t, n, a))
                break

    filtered.sort(key=lambda x: PRIORITY_PARTY.index(x[0]))

    t_out, n_out, a_out = [], [], []
    for item in filtered[:6]:
        t_out.append(item[1])
        n_out.append(item[2])
        a_out.append(item[3])

    return t_out, n_out, a_out


def analyze_and_sort_filings(filings):
    clean = [f for f in filings if f]
    joined = " ".join(clean).lower()

    # -------- Testate Status (robust) --------
    intestate_patterns = [
        "without will",
        "without a will",
        "intestate"
    ]

    if any(p in joined for p in intestate_patterns):
        testate_status = "Intestate"
    elif " will" in joined:  # space prevents matching 'willing'
        testate_status = "Testate"
    else:
        testate_status = ""

    # -------- Filing Priority --------
    ordered = []
    matched = ""

    for phrase in PRIORITY_FILINGS:
        for f in clean:
            if phrase.lower() in f.lower():
                if not matched:
                    matched = f
                ordered.append(f)

    for f in clean:
        if f not in ordered:
            ordered.append(f)

    ordered = ordered[:10] + [""] * (10 - len(ordered))

    return testate_status, matched, ordered


def extract_record_from_details(driver, serial_number):
    """
    Extract all required fields from the Estate Details page
    and return a list (row) matching the Excel header order.
    """

    # Basic case info
    case_no = safe_get_text_by_id(driver, "cpMain_lblCaseNo")
    decedent_name = safe_get_text_by_id(driver, "cpMain_lblCaseName")
    street = safe_get_text_by_id(driver, "cpMain_lblStreetAddress")
    city_state_zip = safe_get_text_by_id(driver, "cpMain_lblCityStateZip")
    dod = safe_get_text_by_id(driver, "cpMain_lblDied")  # print exactly as-is

    # Decedent address as "line1\nline2"
    if street or city_state_zip:
        if street and city_state_zip:
            decedent_address = f"{street}\n{city_state_zip}"
        else:
            decedent_address = street or city_state_zip
    else:
        decedent_address = ""

    # ---------------- PARTIES ----------------
    party_types = []
    party_names = []
    party_addresses = []

    for i in range(20):
        
        p_type = safe_get_text_by_id(driver, f"cpMain_repParty_lblPartyType_{i}")
        if not p_type or not p_type.strip():
            continue

        p_name = safe_get_text_by_id(driver, f"cpMain_repParty_lblParty_{i}")
        addr1 = safe_get_text_by_id(driver, f"cpMain_repParty_lblAddress_{i}")
        addr2 = safe_get_text_by_id(driver, f"cpMain_repParty_lblCityStateZip_{i}")

        if addr1 or addr2:
            if addr1 and addr2:
                full_addr = f"{addr1}\n{addr2}"
            else:
                full_addr = addr1 or addr2
        else:
            full_addr = ""

        party_types.append(p_type)
        party_names.append(p_name)
        party_addresses.append(full_addr)

    # NEW: Filter + priority sort parties
    party_types, party_names, party_addresses = filter_and_sort_parties(
        party_types, party_names, party_addresses
    )


    # Attorney: first Represented By with both name and address; else just name
    attorney_name = ""
    attorney_address = ""
    found_full = False

    for i in range(30):
        for j in range(5):
            name_id = f"cpMain_repParty_repPartyRep_{i}_lblAttorneyName_{j}"
            addr1_id = f"cpMain_repParty_repPartyRep_{i}_lblAttorneyAddress_{j}"
            addr2_id = f"cpMain_repParty_repPartyRep_{i}_lblAttorneyCityStateZip_{j}"

            name = safe_get_text_by_id(driver, name_id)
            if not name:
                continue

            a1 = safe_get_text_by_id(driver, addr1_id)
            a2 = safe_get_text_by_id(driver, addr2_id)

            if a1 or a2:
                if a1 and a2:
                    full_addr = f"{a1}\n{a2}"
                else:
                    full_addr = a1 or a2
            else:
                full_addr = ""

            if name and full_addr and not found_full:
                attorney_name = name
                attorney_address = full_addr
                found_full = True
                break

            if not attorney_name:
                attorney_name = name
                attorney_address = full_addr

        if found_full:
            break

    # Filings
    filing_date = safe_get_text_by_id(driver, "cpMain_repFilings_lblFiledDate_0")

    filings = []
    for i in range(10):
        filing_desc = safe_get_text_by_id(driver, f"cpMain_repFilings_lblFilingTypeDesc_{i}")
        filings.append(filing_desc)
    # NEW: analyze filings
    testate_status, matched_filing, filings = analyze_and_sort_filings(filings)


    # Build row
    row = [
        serial_number,
        "GA",
        case_no,
        decedent_name,
        decedent_address,
        dod,
    ]

    for i in range(6):
        row.append(party_types[i] if i < len(party_types) else "")
        row.append(party_names[i] if i < len(party_names) else "")
        row.append(party_addresses[i] if i < len(party_addresses) else "")

    row.append(attorney_name)
    row.append(attorney_address)
    row.append(filing_date)
    # NEW COLUMNS
    row.append(testate_status)
    row.append(matched_filing)

    for i in range(10):
        row.append(filings[i] if i < len(filings) else "")

    return row


def process_county(driver, wait, county_name, wb, OUTPUT_FILE):
    print(f"\n--- Processing County: {county_name} ---")
    url = "https://www.georgiaprobaterecords.com/Estates/SearchEstates.aspx"

    if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
        ws = wb["Sheet"]
        ws.title = county_name[:31]
    else:
        ws = wb.create_sheet(title=county_name[:31])

    headers = [
        "S.No.", "State", "Case Number", "Decedent Name", "Decedent Address", "DOD",
        "Type 1", "Name 1", "Address 1",
        "Type 2", "Name 2", "Address 2",
        "Type 3", "Name 3", "Address 3",
        "Type 4", "Name 4", "Address 4",
        "Type 5", "Name 5", "Address 5",
        "Type 6", "Name 6", "Address 6",
        "Attorney Name", "Attorney Address", "Filing Date",
        "Testate Status", "Matched Filing",
        "Filing 1", "Filing 2", "Filing 3", "Filing 4", "Filing 5",
        "Filing 6", "Filing 7", "Filing 8", "Filing 9", "Filing 10"
    ]
    ws.append(headers)

    driver.get(url)

    wait.until(EC.presence_of_element_located((By.ID, "ctl00_cpMain_txtFiledStartDate_dateInput")))

    try:
        dropdown = driver.find_element(By.ID, "ctl00_cpMain_ddlCounty")
        dropdown.click()
        time.sleep(1)
        option_xpath = f"//li[text()='{county_name}']"
        option = wait.until(EC.element_to_be_clickable((By.XPATH, option_xpath)))
        option.click()
        time.sleep(1)
    except Exception as e:
        print(f"Could not select county {county_name}: {e}")
        return

    end_date = datetime.now()
    start_date = end_date - timedelta(days=7)
    
    start_date_str = start_date.strftime("%m/%d/%Y")
    end_date_str = end_date.strftime("%m/%d/%Y")

    print(f"Setting dynamically calculated date range: {start_date_str} to {end_date_str}")
    
    start_input = driver.find_element(By.ID, "ctl00_cpMain_txtFiledStartDate_dateInput")
    start_input.clear()
    start_input.send_keys(start_date_str)

    end_input = driver.find_element(By.ID, "ctl00_cpMain_txtFiledEndDate_dateInput")
    end_input.clear()
    end_input.send_keys(end_date_str)

    search_btn = driver.find_element(By.ID, "ctl00_cpMain_btnSearch_input")
    search_btn.click()
    time.sleep(5)

    try:
        wait.until(EC.presence_of_element_located((By.ID, "ctl00_cpMain_rgEstates_ctl00")))
    except Exception as e:
        print(f"Grid not found after search for {county_name}. Expected if no records exist.")
        return

    total_records, total_pages = get_paging_info(driver)
    print(f"Total records (estimated): {total_records}, total pages: {total_pages}")

    rows_now = driver.find_elements(
        By.CSS_SELECTOR,
        "#ctl00_cpMain_rgEstates_ctl00 tbody tr[id^='ctl00_cpMain_rgEstates_ctl00__']"
    )
    if not rows_now:
        wb.save(OUTPUT_FILE)
        print("No records found (grid is empty). Done.")
        return

    main_window = driver.current_window_handle
    serial = 1

    for page_index in range(total_pages):
        wait.until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "#ctl00_cpMain_rgEstates_ctl00 tbody tr[id^='ctl00_cpMain_rgEstates_ctl00__']")
            )
        )

        data_rows = driver.find_elements(
            By.CSS_SELECTOR,
            "#ctl00_cpMain_rgEstates_ctl00 tbody tr[id^='ctl00_cpMain_rgEstates_ctl00__']"
        )
        num_rows = len(data_rows)
        print(f"Page {page_index + 1}/{total_pages}, rows: {num_rows}")

        for row_idx in range(num_rows):
            link_xpath = f"//*[@id='ctl00_cpMain_rgEstates_ctl00__{row_idx}']/td[1]/a"

            wait.until(EC.element_to_be_clickable((By.XPATH, link_xpath)))
            link_el = driver.find_element(By.XPATH, link_xpath)

            href = link_el.get_attribute("href")
            details_url = href if href.startswith("http") else urljoin(driver.current_url, href)

            driver.switch_to.new_window("tab")
            driver.get(details_url)

            wait.until(EC.visibility_of_element_located((By.ID, "cpMain_lblCaseNo")))

            record_row = extract_record_from_details(driver, serial)
            ws.append(record_row)

            wb.save(OUTPUT_FILE)

            print(f"Saved record #{serial}: {record_row[2]} - {record_row[3]}")
            serial += 1

            driver.close()
            driver.switch_to.window(main_window)

        if page_index < total_pages - 1:
            next_btns = driver.find_elements(
                By.CSS_SELECTOR,
                "#ctl00_cpMain_rgEstates_ctl00 > tfoot > tr > td > "
                "table > tbody > tr > td > div.rgWrap.rgArrPart2 > input.rgPageNext"
            )
            if not next_btns:
                print("Next button not found. Stopping pagination.")
                break

            old_first_row = driver.find_element(By.ID, "ctl00_cpMain_rgEstates_ctl00__0")
            next_btns[0].click()
            wait.until(EC.staleness_of(old_first_row))

    wb.save(OUTPUT_FILE)


def main():
    COUNTIES = [
        "Atkinson", "Bacon", "Baldwin", "Barrow", "Bartow", "Ben Hill", "Berrien", "Bibb",
        "Brooks", "Bryan", "Bulloch", "Burke", "Butts", "Carroll", "Catoosa", "Chatham",
        "Chattahoochee", "Chattooga", "Clayton", "Clinch", "Coffee", "Colquitt", "Cook", "Coweta",
        "Crawford", "Crisp", "Dade", "Dawson", "Dougherty", "Douglas", "Effingham", "Elbert",
        "Emanuel", "Fannin", "Fayette", "Floyd", "Franklin", "Gilmer", "Glascock", "Habersham",
        "Hall", "Hancock", "Harris", "Hart", "Henry", "Houston", "Irwin", "Jackson", "Jasper",
        "Jefferson", "Jenkins", "Jones", "Lamar", "Lanier", "Long", "Lumpkin", "Madison",
        "Marion", "McIntosh", "Meriwether", "Monroe", "Morgan", "Murray", "Newton", "Oconee",
        "Oglethorpe", "Paulding", "Pickens", "Pierce", "Pike", "Polk", "Pulaski", "Putnam",
        "Randolph", "Richmond", "Saluda", "Schley", "Spalding", "Talbot", "Taylor", "Terrell",
        "Thomas", "Tift", "Treutlen", "Troup", "Union", "Upson", "Walker", "Walton", "Wheeler",
        "Whitfield", "Wilcox", "Wilkinson", "Worth"
    ]

    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--disable-gpu')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--disable-dev-shm-usage')
    
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    wait = WebDriverWait(driver, 20)

    from openpyxl import Workbook
    wb = Workbook()
    OUTPUT_FILE = f"GA_Daterange_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

    try:
        for c in COUNTIES:
            try:
                driver.current_window_handle
                process_county(driver, wait, c, wb, OUTPUT_FILE)
            except Exception as e:
                print(f"Error processing {c}: {e}. Restarting browser...")
                try:
                    driver.quit()
                except:
                    pass
                chrome_options = Options()
                chrome_options.add_argument('--headless')
                chrome_options.add_argument('--disable-gpu')
                chrome_options.add_argument('--no-sandbox')
                chrome_options.add_argument('--disable-dev-shm-usage')
                driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
                wait = WebDriverWait(driver, 20)
                try:
                    process_county(driver, wait, c, wb, OUTPUT_FILE)
                except Exception as retry_e:
                    print(f"Retry failed for {c}: {retry_e}. Skipping...")
                    
    finally:
        try:
            wb.save(OUTPUT_FILE)
        except Exception as e:
            print("Failed to final save:", e)
        try:
            driver.quit()
        except:
            pass
        print("Scraping completed for all counties. Done.")

if __name__ == "__main__":
    main()
